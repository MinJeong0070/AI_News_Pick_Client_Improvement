from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple
import logging
import pandas as pd
from openpyxl import load_workbook

# =========================
# 경로 설정 (프로젝트 루트 고정)
# =========================
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
LOG_DIR = DATA_DIR / "log"
RES_DIR = PROJECT_ROOT / "resources"

# 디렉터리 보장
LOG_DIR.mkdir(parents=True, exist_ok=True)
(DATA_DIR / "input").mkdir(parents=True, exist_ok=True)
(DATA_DIR / "output").mkdir(parents=True, exist_ok=True)

DEFAULT_INPUT = DATA_DIR / "input" / "default.xlsx"
DEFAULT_OUTPUT = DATA_DIR / "output" / f"output_{datetime.now().strftime('%y%m%d')}.xlsx"


def respath(*names: str) -> Path:
    """resources/ 하위 경로 생성."""
    return RES_DIR.joinpath(*names)


def find_resource(*candidates: str) -> Path:
    """resources/ 아래에서 후보 파일명을 재귀 탐색하여 첫 매치를 반환.
    없으면 RES_DIR/첫 후보 경로를 반환(존재 검사는 호출부에서 처리).
    """
    for name in candidates:
        for p in RES_DIR.rglob(name):
            if p.is_file():
                return p
    return RES_DIR / candidates[0]


# =========================
# 로깅 설정
# =========================
_today = datetime.now().strftime("%y%m%d")
_log_path = LOG_DIR / f"로그_{_today}.txt"

logger = logging.getLogger("preprocess")
logger.setLevel(logging.INFO)
if not logger.handlers:
    _fh = logging.FileHandler(_log_path, encoding="utf-8")
    _sh = logging.StreamHandler()
    _fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    _fh.setFormatter(_fmt)
    _sh.setFormatter(_fmt)
    logger.addHandler(_fh)
    logger.addHandler(_sh)


def log(msg: str) -> None:
    """통일된 로그 출력."""
    logger.info(msg)


# =========================
# 입력 엑셀(하이퍼링크 포함) 로딩
# =========================

def read_excel_with_hyperlinks(file_path: Path | str, sheet: int | str = 0) -> pd.DataFrame:
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"입력 파일 없음: {file_path}")

    wb = load_workbook(filename=str(file_path), data_only=True, read_only=False)
    ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[str(sheet)]

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value is not None else f"COL{idx + 1}"
               for idx, c in enumerate(header_row)]

    rows = []
    for r in ws.iter_rows(min_row=2):
        row_dict = {}
        for col_idx, cell in enumerate(r):
            header = headers[col_idx] if col_idx < len(headers) else f"COL{col_idx + 1}"
            value = cell.value
            hyperlink = cell.hyperlink.target if cell.hyperlink else None
            if header == "게시글제목":
                row_dict[header] = value
                row_dict["게시글URL"] = hyperlink
            else:
                row_dict[header] = value
        rows.append(row_dict)

    df = pd.DataFrame(rows)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# =========================
# 전처리 유틸
# =========================

def preprocess_title(title: Optional[str]) -> Optional[str]:
    """네이버 블로그 검색 꼬리 파라미터 제거."""
    if not isinstance(title, str):
        return title
    return title.split("&keyword=")[0]


def _safe_read_excel_col(path: Path, colname: str) -> list[str]:
    """엑셀 파일에서 특정 컬럼을 안전하게 로드하여 문자열 리스트로 반환.
    파일/컬럼 부재 시 빈 리스트와 경고 로그.
    """
    try:
        if not path.exists():
            log(f"⚠️ 리소스 파일 없음: {path}")
            return []
        df = pd.read_excel(path)
        if colname not in df.columns:
            log(f"⚠️ 컬럼 '{colname}' 없음: {path} / 실제: {list(df.columns)}")
            return []
        return df[colname].dropna().astype(str).str.strip().tolist()
    except Exception as e:
        log(f"⚠️ 리소스 로딩 실패: {path} ({e})")
        return []


# =========================
# 1) 비신탁사(저작권문구/도메인) 필터
# =========================

def filter_untrusted_posts(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    untrusted_file = find_resource(
        "비신탁사_저작권문구+도메인주소.xlsx",
        "비신탁사 저작권문구(언진).xlsx",
    )
    trusted_file = find_resource(
        "매체사_도메인_정보.xlsx",
        "매체사_도메인 정보.xlsx",
    )

    untrusted_copyrights = _safe_read_excel_col(untrusted_file, "저작권 문구")
    untrusted_domains = _safe_read_excel_col(untrusted_file, "도메인")
    trusted_domains = _safe_read_excel_col(trusted_file, "도메인")

    def should_remove(content: object) -> bool:
        s = str(content or "")
        has_untrusted_copyright = any(c and c in s for c in untrusted_copyrights)
        has_untrusted_domain = any(d and d in s for d in untrusted_domains)
        has_trusted_domain = any(d and d in s for d in trusted_domains)
        return (has_untrusted_copyright or has_untrusted_domain) and not has_trusted_domain

    mask = df["게시글내용"].apply(should_remove)
    kept = df[~mask].copy()
    dropped = df[mask].copy()
    log(f"비신탁사 필터링 완료: 유지 {len(kept)}개 / 삭제 {len(dropped)}개")
    return kept, dropped


# =========================
# 2) 비신탁사 매체명 필터
# =========================

def filter_by_untrusted_media_names(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    media_file = find_resource("비신탁사 매체명(전처리).xlsx")
    if not media_file.exists():
        log(f"⚠️ 비신탁사 매체명 파일 없음: {media_file}")
        return df.copy(), df.iloc[0:0].copy()

    try:
        media_names = (
            pd.read_excel(media_file, header=None)
            .iloc[:, 0]
            .dropna()
            .astype(str)
            .tolist()
        )
    except Exception as e:
        log(f"⚠️ 비신탁사 매체명 로딩 실패: {media_file} ({e})")
        return df.copy(), df.iloc[0:0].copy()

    mask = df["게시글내용"].astype(str).apply(lambda x: not any(name in x for name in media_names))
    kept = df[mask].copy()
    dropped = df[~mask].copy()
    log(f"비신탁사 매체명 필터링 완료: 유지 {len(kept)}개 / 삭제 {len(dropped)}개")
    return kept, dropped


# =========================
# 3) 텍스트 규칙 필터 (문장성/만평 등)
# =========================

def filter_empty_image_and_no_da(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """문장성 부족/만평 키워드 등을 기준으로 삭제할 행 마스크 생성."""
    mask = (
            ((~df["게시글제목"].str.contains("다.", regex=False, na=False)) |
             (df["게시글제목"].str.contains("니다.", regex=False, na=False))) &
            ((~df["게시글내용"].str.contains("다.", regex=False, na=False)) |
             (df["게시글내용"].str.contains("니다.", regex=False, na=False))) &
            (~df["게시글제목"].str.contains("만평", regex=False, na=False)) &
            (~df["게시글내용"].str.contains("만평", regex=False, na=False))
    )
    kept = df[~mask].copy()
    dropped = df[mask].copy()
    log(f"텍스트 필터링 완료: 유지 {len(kept)}개 / 삭제 {len(dropped)}개")
    return kept, dropped


# =========================
# 메인 전처리
# =========================

def run_preprocessing(
        input_path: Optional[Path | str] = None,
        output_path: Optional[Path | str] = None,
        stop_event=None,
) -> None:
    input_path = Path(input_path) if input_path else DEFAULT_INPUT
    output_path = Path(output_path) if output_path else DEFAULT_OUTPUT

    log(f"입력 파일: {input_path}")
    log(f"출력 파일: {output_path}")

    df = read_excel_with_hyperlinks(input_path)

    # 기본 정리: 숨은 제어문자 제거, 제목 꼬리 제거
    df = df.replace(to_replace=r"_x000[D|A]_", value=" ", regex=True)
    if "게시글제목" in df.columns:
        df["게시글제목"] = df["게시글제목"].apply(preprocess_title)

    # 제외 ID 로딩
    exclude_ids_path = find_resource("제외 대상 리스트_0925.xlsx", "제외 대상 리스트.xlsx")
    exclude_ids: list[str] = []
    if exclude_ids_path.exists():
        try:
            tmp = pd.read_excel(exclude_ids_path, header=None)
            if "ID" in tmp.columns:
                exclude_ids = tmp["ID"].dropna().astype(str).tolist()
            else:
                exclude_ids = tmp.iloc[:, 0].dropna().astype(str).tolist()
        except Exception as e:
            log(f"⚠️ 제외 대상 로딩 실패: {exclude_ids_path} ({e})")

    # blogId 추출
    from urllib.parse import urlparse, parse_qs

    def extract_blog_id(url: object) -> Optional[str]:
        try:
            q = parse_qs(urlparse(str(url)).query)
            return q.get("blogId", [None])[0]
        except Exception:
            return None

    df["blogId"] = df.get("게시글URL", "").apply(extract_blog_id)
    if exclude_ids:
        before = len(df)
        df = df[~df["blogId"].astype(str).isin(exclude_ids)].copy()
        log(f"제외 ID 적용: {before - len(df)}건 제거 (남은 {len(df)})")

    # 검색어 기반 1차 필터(결측 안전)
    def _has_query(row) -> bool:
        q = str(row.get("검색어", "")).lower()
        title = str(row.get("게시글제목", "")).lower()
        body = str(row.get("게시글내용", "")).lower()
        return (q in title) or (q in body)

    df_q = df[
        df.apply(_has_query, axis=1)
        & (~df.get("게시글내용", "").fillna("").str.contains("신춘문예", case=False))
        & (~df.get("게시글제목", "").fillna("").str.contains("신춘문예", case=False))
        & (~df.get("계정명", "").fillna("").str.contains("뽐뿌뉴스", case=False))
        ].copy()

    log(f"검색어/예외어 필터 적용: 유지 {len(df_q)}건 / 삭제 {len(df) - len(df_q)}건")

    # 비었으면 빈 파일 저장 후 종료
    if df_q.empty:
        log("⚠️ 검색어 기반 필터 결과가 비었습니다. 빈 파일 저장 후 종료.")
        df_q.head(0).to_excel(output_path, index=False)
        log(f"✅ 전처리 완료. 저장: {output_path}")
        return
    if stop_event and stop_event.is_set():
        log("🛑 사용자 중단 요청, 작업 중단")
        return

    # 2차: 비신탁사(저작권/도메인)
    df1, _ = filter_untrusted_posts(df_q)
    if df1.empty:
        log("⚠️ 비신탁사 필터 결과가 비었습니다. 빈 파일 저장 후 종료.")
        df1.head(0).to_excel(output_path, index=False)
        log(f"✅ 전처리 완료. 저장: {output_path}")
        return
    if stop_event and stop_event.is_set():
        log("🛑 사용자 중단 요청, 작업 중단")
        return

    # 3차: 비신탁사 매체명
    df2, _ = filter_by_untrusted_media_names(df1)
    if df2.empty:
        log("⚠️ 매체명 필터 결과가 비었습니다. 빈 파일 저장 후 종료.")
        df2.head(0).to_excel(output_path, index=False)
        log(f"✅ 전처리 완료. 저장: {output_path}")
        return
    if stop_event and stop_event.is_set():
        log("🛑 사용자 중단 요청, 작업 중단")
        return

    # 4차: 텍스트 규칙 필터(문장성/만평 등)
    df3, _ = filter_empty_image_and_no_da(df2)
    if df3.empty:
        log("⚠️ 텍스트 필터 결과가 비었습니다. 빈 파일 저장 후 종료.")
        df3.head(0).to_excel(output_path, index=False)
        log(f"✅ 전처리 완료. 저장: {output_path}")
        return
    if stop_event and stop_event.is_set():
        log("🛑 사용자 중단 요청, 작업 중단")
        return

    # 저장
    df3.drop(columns=["blogId"], inplace=True, errors="ignore")
    df3.to_excel(output_path, index=False)
    log(f"✅ 전처리 완료. 저장: {output_path}")


if __name__ == "__main__":
    run_preprocessing()
